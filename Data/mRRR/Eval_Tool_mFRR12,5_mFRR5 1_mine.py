#Marek Bistricky, Nano Energies
#March 2024
#This code serves as an evaluation tool of certification measurements for ancillary services @ Nano Energies

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import math
import statistics as stc
import openpyxl
from openpyxl import load_workbook

#NECESSARY INPUT------------------------------------------------------------------

#enter !EXCEL LINE! where the 2nd command took place - t2 limit vertical line - 2nd command, DOWN for mFRR+ and UP for mFRR-
#ramp_second_command_t2 = 1995; #B
#ramp_second_command_t2 = 2072; #A
#ramp_second_command_t2 = 2102
#ramp_second_command_t2 = 2032
ramp_second_command_t2 = 2012
#ramp_second_command_t2 = 1562 #mFRR5

excel_name = 'DES_data_DES_ČEPS CERT_CEPS_mFRR125_A_2024-09-10_Dětenice_1MW.xlsx'
#excel_name = 'DES_data_DES_CEPS CERT_CEPS_mFRR125_B_1,8MW.xlsx'
#excel_name = 'DES_data_DES_CEPS CERT_CEPS_mFRR125_Aminus_1,8MW.xlsx'
#excel_name = 'AB02_DES_data_DES_AB02-mFRR12,5B_1_MW (2).xlsx'
#excel_name = 'DES_data_DES_CEPS CERT_CEPS_mFRR125_B_Jipocar1MW.xlsx'

#tolerance_coef = 0.14 #old tolerance coefficient not valid since 1.1.2023
tolerance_coef = 0.2

ceps_product = 0 #0 for mfrr12.5  || 1 for mfrr5
battery = 0 #1 if battery is certified, 0 otherwise

if ceps_product==0:
    ramp_length = 750
    ramp_second_command_t2 = 2012
elif ceps_product==1:
    ramp_length = 300
    ramp_second_command_t2 = 1562
#----------------------------------------------------------------------------------

excel_source = pd.read_excel(excel_name)
support_excel_load = load_workbook(excel_name, data_only=True)

#print(excel_source.shape)
#print(excel_source.columns)

P_skut = excel_source['PSKUT_AB (MW)']
exact_date_time = excel_source['Date']
test_time = excel_source['Time']
P_cil = excel_source['mFRRZADDA (MW)'][1]
P_dg = excel_source['PDG_AB (MW)'][1]
G_av = excel_source['GV_AB']
if battery == 1:
    Soc = excel_source['SoC (%)']

#print(P_cil+P_dg)
#print(round(P_skut[804],2))

sheet_name = support_excel_load.sheetnames[0]
sh = support_excel_load[sheet_name]
first_column = sh['A']
color_in_hex = sh['A%s'%(ramp_second_command_t2)].fill.start_color.index
#print('HEX =',color_in_hex)
#print('RGB =', tuple(int(color_in_hex[i:i+2], 16) for i in (0, 2, 4))) # Color in RGB

# getting the tolerance 
P_dov = min(5, tolerance_coef*P_cil)
print("P_dov=",P_dov)

#12.5min time interval from second command
t2_plus_12_5 = ramp_second_command_t2+ramp_length;


#getting the t0 limit vertical line / beginning of the whole certification process when t=1
for x in range(excel_source.shape[0]-1):
    if math.isnan(P_skut[x])==False and math.isnan(P_skut[x+1])==False:
        if test_time[x]==1:
            beginning_t0 = x+1

value_found = 0

#getting tf - first change of power
for x in range(excel_source.shape[0]-1):
    if math.isnan(P_skut[x])==False and math.isnan(P_skut[x+1])==False:
        if x>2 and abs(P_skut[x]-P_skut[x-1])>0 and value_found==0 and abs(G_av[x]-G_av[x-1])>0: #and P_dg+P_cil==0:
            prifazovani = x+1
            value_found=value_found+1
        elif value_found<1 :
          prifazovani = 1000 # tady vymysli jak to udelat jinak pro GreenGas

#getting the t1 limit vertical line - reached P_cil
value_found=0
for x in range(excel_source.shape[0]-1):
    if math.isnan(P_skut[x])==False and math.isnan(P_skut[x+1])==False:
        if round(P_skut[x],2)==(round(P_dg+P_cil,2)) and value_found==0 and test_time[x] < ramp_length+60: #changed == for <= and added abs for Pskut
            reached_P_cil_t1 = x+1
            value_found=value_found+1
            print(round(P_skut[x],2))
            print(test_time[x])
        if value_found<1:
            reached_P_cil_t1 = ramp_length+61

print(reached_P_cil_t1)

value_found=0

# 12.5min time interval from start command
for x in range(excel_source.shape[0]-1):
    if math.isnan(P_skut[x])==False and math.isnan(P_skut[x+1])==False:
        if test_time[x]==ramp_length+1:
            t0_plus_12_5 = x+1

#getting the t2 limit vertical line - second command
#for x in range(excel_source.shape[0]-1):
 #   if math.isnan(P_skut[x])==False and math.isnan(P_skut[x+1])==False:
  #      color_in_hex2 = first_column[x].fill.start_color.index
   #     if color_in_hex2 == 'FFFF0000' and value_found==0: #and round(P_skut[x],2) == (P_dg + P_cil):
    #        trial_time = x+1
     #       value_found = value_found+1

value_found=0

#print(trial_time)

#getting the t3 limit vertical line - reached P_dg
if P_dg==0: #test A
    for x in range(excel_source.shape[0]-1):
        if math.isnan(P_skut[x])==False and math.isnan(P_skut[x+1])==False:
            if test_time[x]>ramp_second_command_t2 and P_skut[x]==P_dg and value_found==0 and abs(G_av[x]-G_av[x-1])>0:
                reached_P_dg_t3 = x+1
                value_found=value_found+1
elif P_dg < 0: #consumption and Pskut needs to be lover or equal than Pdg
    for x in range(excel_source.shape[0]-1):
        if math.isnan(P_skut[x])==False and math.isnan(P_skut[x+1])==False:
            if test_time[x]>ramp_second_command_t2 and round(P_skut[x],2)<=P_dg and value_found==0 :
                reached_P_dg_t3 = x+1
                value_found=value_found+1
            elif test_time[x]==ramp_second_command_t2-60+ramp_length:
                reached_P_dg_t3 = x+1
                value_found = value_found+1
else: #Pdg > 0
    for x in range(excel_source.shape[0]-1):
        if math.isnan(P_skut[x])==False and math.isnan(P_skut[x+1])==False:
            if test_time[x]>ramp_second_command_t2 and test_time[x] < ramp_second_command_t2+ramp_length+1 and round(P_skut[x],2)==P_dg and value_found==0 :
                reached_P_dg_t3 = x+1
                value_found=value_found+1

            if value_found<1:
                reached_P_dg_t3 = ramp_second_command_t2+ramp_length

#print(test_time[2769])
#print(round(P_skut[2769],2))
print(P_dg)


#getting the t4 limit vertical limit - last 20min in test B
if P_dg==0 :
    t4 = "Not relevant for Test A"
    print("\n Povel k aktivaci: " , exact_date_time[beginning_t0-1] , "\n Prifazovani: " , exact_date_time[prifazovani-1] , "" , test_time[prifazovani-1] , "\n t1: " , exact_date_time[reached_P_cil_t1-1] , "" , test_time[reached_P_cil_t1-1] ,
        "\n Povel k deaktivaci: " , exact_date_time[ramp_second_command_t2-2] , "" , test_time[ramp_second_command_t2-2] , "\n t3: " , exact_date_time[reached_P_dg_t3-1], "" , test_time[reached_P_dg_t3-1])
else:
    t4 = t2_plus_12_5+(20*60)
    print("\n Povel k aktivaci: " , exact_date_time[beginning_t0-1] , "\n Prifazovani: " ,exact_date_time[prifazovani-1] , "" , test_time[prifazovani-1] , "\n t1: " , exact_date_time[reached_P_cil_t1-1] , "" , test_time[reached_P_cil_t1-1] ,
        "\n Povel k deaktivaci: " , exact_date_time[ramp_second_command_t2-2] , "" , test_time[ramp_second_command_t2-2] , "\n t3: " , exact_date_time[reached_P_dg_t3-1] , "" , test_time[reached_P_dg_t3-1] , "\n t4: " , exact_date_time[t4-1] , "" , test_time[t4-1])

value_found=0

print("\n P_{dov} = ", P_dov , " MW")

print("\n These are lines in Excel with CEPS limits from codex to be but into the final report: \n t0:" , beginning_t0+1 ,"\n tf:" , prifazovani+1 , " \n t1:" , reached_P_cil_t1+1 , " \n t2:" , ramp_second_command_t2 , " \n t3:" , reached_P_dg_t3+1 , " \n t4:" , t4)

# getting activation and deactiovation times

t_akt = reached_P_cil_t1-beginning_t0

t_deakt = reached_P_dg_t3-ramp_second_command_t2

#getting tolerance curves

P_lim_plus = [];
P_lim_minus =[];

if P_dg==0: #mFRR+ and TEST A

    #deviation and standard deviation in <t1; t2> ----------------------------------------------------
    data_t1_t2 = P_skut[reached_P_cil_t1:ramp_second_command_t2]
    mean_data_t1_t2 = stc.mean(P_skut[reached_P_cil_t1:ramp_second_command_t2])
    diff = []
    for x in data_t1_t2:    #range(ramp_second_command_t2-reached_P_cil_t1)
        diff.append(abs(P_cil - x))

    avg_deviation_A = (sum(diff)/(len(data_t1_t2))) #A

    if avg_deviation_A < 0.25*P_dov:
        res = True
    else:
        res = False
    
    standard_deviation_Sigma = stc.stdev(data_t1_t2)

    if standard_deviation_Sigma < P_dov:
        res2 = True
    else:
        res2 = False

    print("\n These are average (A) and standard (Sigma) deviations in <t1; t2> for Test A: \n A:" , avg_deviation_A , "        A < 0.25*P_dov " , 0.25*P_dov , " is " , res , 
            " \n Sigma:" , standard_deviation_Sigma, "           Sigma < P_dov is " , res2)

    #--------------------------------------------------------------------------------------------------

    for x in range(excel_source.shape[0]):
        if x < beginning_t0  and math.isnan(test_time[x])==True:
            P_lim_plus.append(0+P_dov)
            P_lim_minus.append(0-P_dov)
        elif x < t0_plus_12_5:
            #P_lim_plus.append((((P_cil+(P_dov)-(P_dov))/750)*(test_time[x]-1))+(P_dov)) #old limits
            P_lim_plus.append(P_cil+P_dov)
            P_lim_minus.append(0-P_dov)
        elif x >= t0_plus_12_5 and x <= ramp_second_command_t2:
            P_lim_plus.append(P_lim_plus[x-1])
            P_lim_minus.append(P_cil-P_dov)
        elif x > ramp_second_command_t2 and x < t2_plus_12_5:
            P_lim_plus.append(P_lim_plus[x-1])
            #P_lim_minus.append((((P_cil-(P_dov)+(P_dov))/750)*(-test_time[x]+test_time[t2_plus_12_5]))-(P_dov))  #old limits
            P_lim_minus.append(0-P_dov)
        else:
            P_lim_plus.append(0+P_dov)
            P_lim_minus.append(0-P_dov)

    #If the test is done for consumption (light, etc.), the limit curves need to be switched
    if (P_dg == 0 and P_cil < 0) or (P_dg > 0):
        help_variable = P_lim_minus
        P_lim_minus = P_lim_plus
        P_lim_plus = help_variable


else: #mFRR- and TEST B

    #deviation and standard deviation in <t1; t2> ----------------------------------------------------

    data_t1_t2 = P_skut[reached_P_cil_t1:ramp_second_command_t2]
    mean_data_t1_t2 = stc.mean(P_skut[reached_P_cil_t1:ramp_second_command_t2])
    desired_value = P_dg+P_cil
    diff = []
    for x in data_t1_t2:
        diff.append(abs(desired_value - x))
    
    avg_deviation_A = sum(diff)/(len(data_t1_t2)) #A

    if avg_deviation_A < 0.25*P_dov:
        res = True
    else:
        res = False
    
    standard_deviation_Sigma = stc.stdev(data_t1_t2)

    if standard_deviation_Sigma < P_dov:
        res2 = True
    else:
        res2 = False

    #deviation and standard deviation in <t3; t4> ----------------------------------------------------
    data_t3_t4 = P_skut[reached_P_dg_t3:t4]
    mean_data_t3_t4 = stc.mean(P_skut[reached_P_dg_t3:t4])
    diff2 = []
    for x in data_t3_t4:
        diff2.append(abs(P_dg - x))

    avg_deviation_A2 = (sum(diff2)/(len(data_t3_t4))) #A2

    if avg_deviation_A2 < 0.25*P_dov:
        res3 = True
    else:
        res3 = False
    
    standard_deviation_Sigma2 = stc.stdev(data_t3_t4)

    if standard_deviation_Sigma2 < P_dov:
        res4 = True
    else:
        res4 = False

    print("\n These are average (A, A2) and standard (Sigma, Sigma2) deviations for Test B in <t1; t2>: \n A:" , avg_deviation_A, "           A < 0.25*P_dov is " , 0.25*P_dov , "" , res ,
            "\n Sigma:" , standard_deviation_Sigma , "     Sigma < P_dov is " , res2 , " \n and in <t3; t4> \n A2:" , avg_deviation_A2 , "        A2 < 0.25*P_dov is " , res3 , 
            " \n Sigma2:" , standard_deviation_Sigma2, "           Sigma2 < P_dov is " , res4)

    #--------------------------------------------------------------------------------------------------

    if P_dg < 0:
        for x in range(excel_source.shape[0]):
            if x < beginning_t0 and math.isnan(test_time[x])==True:
                P_lim_minus.append(P_dg+abs(P_dov))
                P_lim_plus.append(P_dg-abs(P_dov))
            elif x < t0_plus_12_5:
                P_lim_plus.append(P_dg-P_dov)
                P_lim_minus.append(P_dg+P_cil+P_dov)
            elif x >= t0_plus_12_5 and x <= ramp_second_command_t2:
                P_lim_minus.append((P_dg + P_cil)+abs(P_dov))
                P_lim_plus.append((P_dg + P_cil)-abs(P_dov))
            elif x > ramp_second_command_t2 and x < t2_plus_12_5:
                P_lim_plus.append(P_dg-P_dov)
                P_lim_minus.append(P_dg+P_cil+P_dov)
            else:
                P_lim_minus.append(P_dg+abs(P_dov))
                P_lim_plus.append(P_dg-abs(P_dov))
    elif P_dg > 0:
        for x in range(excel_source.shape[0]):
            if x < beginning_t0 and math.isnan(test_time[x])==True:
                P_lim_minus.append(P_dg+abs(P_dov))
                P_lim_plus.append(P_dg-abs(P_dov))
            elif x < t0_plus_12_5:
                P_lim_minus.append(P_dg-P_dov)
                P_lim_plus.append(P_dg+P_cil+P_dov)
            elif x >= t0_plus_12_5 and x <= ramp_second_command_t2:
                P_lim_plus.append((P_dg + P_cil)+P_dov)
                P_lim_minus.append((P_dg + P_cil)-P_dov)
            elif x > ramp_second_command_t2 and x < t2_plus_12_5:
                P_lim_minus.append(P_dg-P_dov)
                P_lim_plus.append(P_dg+P_cil+P_dov)
            else:
                P_lim_minus.append(P_dg+abs(P_dov))
                P_lim_plus.append(P_dg-abs(P_dov))


value_found = 0

for x in range(excel_source.shape[0]):
    if P_dg == 0: #mFRR+ and Test A
        if round(P_skut[x],2)>=P_cil and value_found == 0:
            t0_plus_12_5 = x+1
            value_found = 1
    else: #mFRR- and Test B
        if round(P_skut[x],2)==P_cil and value_found == 0:
            t0_plus_12_5 = x+1
            value_found = 1

#If the test is done for consumption (light, etc.), the limit curves need to be switched
if (P_dg == 0 and P_cil < 0) or (P_dg > 0):
   help_variable = P_lim_minus
   #P_lim_minus = P_lim_plus
   #P_lim_plus = help_variable

mimo_hodnoty = 0
mezi_hodnoty = 0

#98% limit verification
for x in range(60, excel_source.shape[0]):
    if P_dg == 0:
        if P_skut[x] > P_lim_plus[x] or P_skut[x] < P_lim_minus[x]: #mFRR+ and Test A
            mimo_hodnoty = mimo_hodnoty+1
        else:
            mezi_hodnoty = mezi_hodnoty+1
    else:
        if P_skut[x] < P_lim_plus[x] or P_skut[x] > P_lim_minus[x]: #Test B
            mimo_hodnoty = mimo_hodnoty+1
        else:
            mezi_hodnoty = mezi_hodnoty+1

#check if mimohodnoty and mezihodnoty match with total sample
if mezi_hodnoty+mimo_hodnoty == excel_source.shape[0]-60:
    print("Match is True")
    procento_mezi = mezi_hodnoty/test_time[excel_source.shape[0]-1]
    print(mezi_hodnoty)
    print(mimo_hodnoty)
    print(procento_mezi)
    print(test_time[excel_source.shape[0]-1])
    print(excel_source.shape[0])

#PLOTS--------------------------------------------------------------------------------------------

font_size = 15

if P_dg==0: #mFRR+ and TEST A

    fig, ax = plt.subplots()

    font = {'family' : 'DejaVu Sans',
        'weight' : 'bold',
        'size'   : font_size}

    font2 = {'family' : 'DejaVu Sans',
        'weight' : 'normal',
        'size'   : font_size}

    plt.plot(P_lim_plus, color = 'orange', linewidth = '2', label = r'$ P_{lim,+} $')
    plt.plot(P_lim_minus,  color = 'orange', linewidth = '2', label = r'$ P_{lim,-} $')
    plt.vlines(x = beginning_t0, ymin = min(P_lim_minus)-0.6, ymax = max(P_lim_plus)+0.35, 
                color = '#22A708', ls='--', label = 't0_lim') #t0 vertical line - beginning of the certification process
    plt.text(beginning_t0-10, max(P_lim_plus)+0.37, r'$ t_{0} $')
    plt.vlines(x = prifazovani, ymin = min(P_lim_minus)-0.6, ymax = max(P_lim_plus)+0.35, 
                color = '#22A708', ls='--', label = 'tf_lim') #tf vertical line - first change of power
    plt.text(prifazovani-10, max(P_lim_plus)+0.37, r'$ t_{f} $')
    plt.vlines(x = reached_P_cil_t1, ymin = min(P_lim_minus)-0.6, ymax = max(P_lim_plus)+0.35, 
                color = '#22A708', label = 't1_lim', ls = '--') #t1 vertical line - where the 12.5min ramp ends
    plt.text(reached_P_cil_t1-10, max(P_lim_plus)+0.37, r'$ t_{1} $')
    plt.vlines(x = ramp_second_command_t2, ymin = min(P_lim_minus)-0.6, ymax = max(P_lim_plus)+0.35,
                color = '#22A708', label = 't2_lim', ls = '--') #t2 vertical line - where the second command takes place
    plt.text(ramp_second_command_t2-10, max(P_lim_plus)+0.37, r'$ t_{2} $')
    plt.vlines(x = reached_P_dg_t3, ymin = min(P_lim_minus)-0.6, ymax = max(P_lim_plus)+0.35,
                color = '#22A708', label = 't3_lim', ls = '--') #t3 vertical line - where the 12.5min ramp down ends
    plt.text(reached_P_dg_t3-10, max(P_lim_plus)+0.37, r'$ t_{3} $')
    plt.axhline(y = P_dg, color = 'y', label = r'$ P_{DG} $')
    plt.text(P_dg+1500, 0.1, r'$ P_{} = {} (MW) $'.format('{DG}', math.floor(P_dg)))
    plt.axhline(y = P_cil, color = 'r', label = r'$ P_{cíl} $')
    plt.text(250, P_cil+0.1, r'$ P_{} = {} (MW) $'.format('{cíl}', (round(P_cil,2))))
    if ceps_product==0:
        plt.annotate(text='', xy=(reached_P_cil_t1,min(P_lim_minus)-0.6), xytext=(beginning_t0,min(P_lim_minus)-0.6), arrowprops=dict(arrowstyle='<->')) #line between t0 and t1
        plt.text((beginning_t0+(t_akt/3))+5, min(P_lim_minus)-0.5, (r'$ t_{AKTmFRR12,5} = %s (s) $'%(t_akt)))
        plt.annotate(text='', xy=(reached_P_dg_t3,min(P_lim_minus)-0.6), xytext=(ramp_second_command_t2,min(P_lim_minus)-0.6), arrowprops=dict(arrowstyle='<->')) #line between t2 and t3
        plt.text((ramp_second_command_t2+(t_deakt/3))+5, min(P_lim_minus)-0.5, (r'$ t_{DEAKTmFRR12,5} = %s (s) $'%(t_deakt)))
        plt.title(r'$ Průběh\ testu\ mFRR_{12.5A}\ -\ grafy:\ P_{cíl}\ =\ P_{DG}\ ±\ mFRR_{ZAD}\ =\ f(t);\ PSKUT\ =\ f(t) $', **font)
    elif ceps_product == 1:
        plt.annotate(text='', xy=(reached_P_cil_t1,min(P_lim_minus)-0.6), xytext=(beginning_t0,min(P_lim_minus)-0.6), arrowprops=dict(arrowstyle='<->')) #line between t0 and t1
        plt.text((beginning_t0+(t_akt/3))+5, min(P_lim_minus)-0.5, (r'$ t_{AKTmFRR5} = %s (s) $'%(t_akt)))
        plt.annotate(text='', xy=(reached_P_dg_t3,min(P_lim_minus)-0.6), xytext=(ramp_second_command_t2,min(P_lim_minus)-0.6), arrowprops=dict(arrowstyle='<->')) #line between t2 and t3
        plt.text((ramp_second_command_t2+(t_deakt/3))+5, min(P_lim_minus)-0.5, (r'$ t_{DEAKTmFRR5} = %s (s) $'%(t_deakt)))
        plt.title(r'$ Průběh\ testu\ mFRR_{5A}\ -\ grafy:\ P_{cíl}\ =\ P_{DG}\ ±\ mFRR_{ZAD}\ =\ f(t);\ PSKUT\ =\ f(t) $', **font)
        
    plt.xlabel(r'$ Čas\ [s]$', **font2)
    plt.ylabel(r'$Výkon\ [MW]$', **font2)

    plt.xlim(0, excel_source.shape[0]+100)
    plt.ylim(min(P_lim_minus)-0.8, max(P_lim_plus)+0.6)

    plt.plot(P_skut, color = 'b', linewidth = '2', label =r'$ P_{SKUT,AB}\ (MW) $')

    # Shrink current axis's height by 10% on the bottom
    box = ax.get_position()
    ax.set_position([box.x0, box.y0 + box.height * 0.05, box.width, box.height * 0.95])

    # Put a legend below current axis
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.1), fancybox=True, shadow=True, ncol=5)

    plt.xticks(np.arange(0, excel_source.shape[0]+100, step=200), **font2)
    plt.yticks(np.arange(round( min(P_lim_minus)-0.8, 0), round(max(P_lim_plus)+0.6, 0), step=0.5), **font2)

    plt.rc('font', **font)

    plt.show()

else: #mFRR- and TEST B

    fig, ax = plt.subplots()

    font_size = 12
    font = {'family': 'DejaVu Sans', 'weight': 'bold', 'size': font_size}
    font2 = {'family': 'DejaVu Sans', 'weight': 'normal', 'size': font_size}

    fig, ax = plt.subplots()

    plt.plot(P_lim_plus, color='orange', linewidth=2, label=r'$ P_{lim,+} $')
    plt.plot(P_lim_minus, color='orange', linewidth=2, label=r'$ P_{lim,-} $')
    plt.vlines(x=beginning_t0, ymin=-1, ymax=1, color='#22A708', ls='--', label='t0_lim')
    plt.text(beginning_t0 - 10, 1.05, r'$ t_{0} $', fontsize=font_size)
    plt.vlines(x=reached_P_cil_t1, ymin=-1, ymax=1, color='#22A708', label='t1_lim', ls='--')
    plt.text(reached_P_cil_t1 - 10, 1.05, r'$ t_{1} $', fontsize=font_size)
    plt.vlines(x=ramp_second_command_t2, ymin=-1, ymax=1, color='#22A708', label='t2_lim', ls='--')
    plt.text(ramp_second_command_t2 - 10, 1.05, r'$ t_{2} $', fontsize=font_size)
    plt.vlines(x=reached_P_dg_t3, ymin=-1, ymax=1, color='#22A708', label='t3_lim', ls='--')
    plt.text(reached_P_dg_t3 - 10, 1.05, r'$ t_{3} $', fontsize=font_size)
    plt.vlines(x=t4, ymin=-1, ymax=1, color='#22A708', label='t4_lim', ls='--')
    plt.text(t4 - 10, 1.05, r'$ t_{4} $', fontsize=font_size)
    plt.axhline(y=P_cil, color='y', label=r'$ P_{cíl} $')
    plt.text(reached_P_dg_t3 + 500, P_cil + 0.1, r'$ P_{} = {} (MW) $'.format('{cíl}', round(P_cil, 2)),
             fontsize=font_size)
    plt.axhline(y=P_dg, color='r', label=r'$ P_{DG} $')
    plt.text(reached_P_dg_t3 + 500, P_dg + 0.1, r'$ P_{} = {} (MW) $'.format('{DG}', round(P_dg, 2)),
             fontsize=font_size)
    plt.plot(P_skut, color='b', linewidth=2, label=r'$ P_{SKUT,AB}\ (MW) $')

    plt.annotate('', xy=(reached_P_cil_t1, -1), xytext=(beginning_t0, -1), arrowprops=dict(arrowstyle='<->'))
    plt.text((beginning_t0 + t_akt / 3) + 5, -0.95, r'$ t_{AKTmFRR12,5} = %s (s) $' % (t_akt), fontsize=font_size)
    plt.annotate('', xy=(reached_P_dg_t3, -1), xytext=(ramp_second_command_t2, -1), arrowprops=dict(arrowstyle='<->'))
    plt.text((ramp_second_command_t2 + t_deakt / 3) + 5, -0.95, r'$ t_{DEAKTmFRR12,5} = %s (s) $' % (t_deakt),
             fontsize=font_size)

    plt.title(
        r'$ Průběh\ testu\ mFRR_{12.5B}\ -\ grafy:\ P_{cíl}\ =\ P_{DG}\ ±\ mFRR_{ZAD}\ =\ f(t);\ PSKUT\ =\ f(t) $',
        **font)
    plt.xlabel(r'$ Čas\ [s]$', **font2)
    plt.ylabel(r'$Výkon\ [MW]$', **font2)
    plt.xlim(0, 1500)
    plt.ylim(0.1, 1)

    # Adjust the legend and ticks
    box = ax.get_position()
    ax.set_position([box.x0, box.y0 + box.height * 0.05, box.width, box.height * 0.95])
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.1), fancybox=True, shadow=True, ncol=5)
    plt.xticks(np.arange(0, 1600, step=200), **font2)
    plt.yticks(np.arange(0.1, 1.1, step=0.2), **font2)

    plt.show()
#--------------------------------------------------------------------------------------------------------------

#ideal plots - certification plan
start_of_plot = 500

font3 = {'family' : 'DejaVu Sans',
        'weight' : 'normal',
        'size'   : font_size-5}

if P_dg==0:

    fig,ax = plt.subplots()

    plt.xlim(0, 3000+ramp_length)
    plt.ylim(min(P_lim_minus)-0.8, max(P_lim_plus)+0.6)
    point1 = [0 , 0]
    point2 = [start_of_plot , 0]
    point_tf = [start_of_plot+150,0]
    point3 = [start_of_plot+ramp_length, P_cil]
    point4 = [start_of_plot+ramp_length+20*60, P_cil]
    point5 = [start_of_plot+ramp_length+20*60+ramp_length, 0]
    point6 = [start_of_plot+ramp_length+20*60+ramp_length+start_of_plot, 0]

    point7= [start_of_plot,P_cil]
    point8=[start_of_plot+ramp_length+20*60, 0]

    x_values = [point1[0], point2[0], point_tf[0], point3[0], point4[0], point5[0], point6[0]]
    y_values = [point1[1], point2[1],  point_tf[1], point3[1], point4[1], point5[1], point6[1]]

    plt.hlines(y=0, xmin=0, xmax=3800, color='#000000')
    plt.text(50,0.07, r'$ P_{} = {} (MW) $'.format('{DG}', (P_dg)))
    plt.hlines(y=P_cil, xmin=0, xmax=3800, color='#000000')
    if ceps_product==0:
        plt.text(50,P_cil+0.07, r'$ mFRR_{12,5A} = %s (MW) $'%(round(P_cil,2)))
    elif ceps_product==1:
        plt.text(50,P_cil+0.07, r'$ mFRR_{5A} = %s (MW) $'%(round(P_cil,2)))

    plt.vlines(x=point2[0], ymin= min(P_lim_minus)-0.6, ymax = max(P_lim_plus)+0.35, ls = '--', color='#000000')
    plt.vlines(x=point_tf[0], ymin= min(P_lim_minus)-0.6, ymax = max(P_lim_plus)+0.35, ls = '--', color='#000000')
    plt.vlines(x=point3[0], ymin= min(P_lim_minus)-0.6, ymax = max(P_lim_plus)+0.35, ls = '--', color='#000000')
    plt.vlines(x=point4[0], ymin= min(P_lim_minus)-0.6, ymax = max(P_lim_plus)+0.35, ls = '--', color='#000000')
    plt.vlines(x=point5[0], ymin= min(P_lim_minus)-0.6, ymax = max(P_lim_plus)+0.35, ls = '--', color='#000000')
    plt.xticks([point2[0], point_tf[0], point3[0], point4[0], point5[0]], [r'$t_{0}$', r'$t_{f} $', r'$t_{1}$', r'$t_{2}$', r'$t_{3}$'])
    plt.yticks([0, round(P_cil/2,2), round(P_cil,2)])
    if ceps_product==0:
        plt.annotate(text='', xy=(point3[0],max(P_skut+0.6)), xytext=(point2[0],max(P_skut+0.6)), arrowprops=dict(arrowstyle='<->'))
        plt.text(point2[0]+40, max(P_skut+0.7), r'$ t_{AKTmFRR12,5} <= 12,5 (min) $', **font) 
        plt.annotate(text='', xy=(point4[0],max(P_skut+0.6)), xytext=(point3[0],max(P_skut+0.6)), arrowprops=dict(arrowstyle='<->')) 
        plt.text(point3[0]+400, max(P_skut+0.7), r'$ t_{u} = 20 (min) $', **font)
        plt.annotate(text='', xy=(point5[0],max(P_skut+0.6)), xytext=(point4[0],max(P_skut+0.6)), arrowprops=dict(arrowstyle='<->'))
        plt.text(point4[0]+40, max(P_skut+0.7), r'$ t_{DEAKTmFRR12,5} <= 12,5 (min) $', **font)
        plt.text(point5[0]+100, P_cil/2, (r'$ t_{0} $ = čas vydání povelu k aktivaci $ mFRR_{12,5} $'), **font3)
        plt.text(point5[0]+100, (P_cil/2)-0.15*(P_cil/4), (r'$ t_{f} $ = čas přifázování bloku k ES '), **font3)
        plt.text(point5[0]+100, (P_cil/2)-0.3*(P_cil/4), (r'$ t_{1} $ = čas dosažení certifikované hodnoty $ mFRR_{12,5} $'), **font3)
        plt.text(point5[0]+100, (P_cil/2)-0.45*(P_cil/4), (r'$ t_{2} $ = čas vydání povelu k deaktivaci $ mFRR_{12,5} $'), **font3) 
    elif ceps_product==1:
        plt.annotate(text='', xy=(point3[0],max(P_skut+0.6)), xytext=(point2[0],max(P_skut+0.6)), arrowprops=dict(arrowstyle='<->'))
        plt.text(point2[0]-220, max(P_skut+0.7), r'$ t_{AKTmFRR5} <= 5 (min) $', **font) 
        plt.annotate(text='', xy=(point4[0],max(P_skut+0.6)), xytext=(point3[0],max(P_skut+0.6)), arrowprops=dict(arrowstyle='<->')) 
        plt.text(point3[0]+400, max(P_skut+0.7), r'$ t_{u} = 20 (min) $', **font)
        plt.annotate(text='', xy=(point5[0],max(P_skut+0.6)), xytext=(point4[0],max(P_skut+0.6)), arrowprops=dict(arrowstyle='<->'))
        plt.text(point4[0]+40, max(P_skut+0.7), r'$ t_{DEAKTmFRR5} <= 5 (min) $', **font)
        plt.text(point5[0]+100, P_cil/2, (r'$ t_{0} $ = čas vydání povelu k aktivaci $ mFRR_{5} $'), **font3)
        plt.text(point5[0]+100, (P_cil/2)-0.15*(P_cil/4), (r'$ t_{f} $ = čas přifázování bloku k ES '), **font3)
        plt.text(point5[0]+100, (P_cil/2)-0.3*(P_cil/4), (r'$ t_{1} $ = čas dosažení certifikované hodnoty $ mFRR_{5} $'), **font3)
        plt.text(point5[0]+100, (P_cil/2)-0.45*(P_cil/4), (r'$ t_{2} $ = čas vydání povelu k deaktivaci $ mFRR_{5} $'), **font3) 
    
    plt.text(point5[0]+100, (P_cil/2)-0.6*(P_cil/4), (r'$ t_{3} $ = čas odepnutí bloku od ES '), **font3)    

    x_values2 = [point1[0], point2[0], point7[0], point4[0], point8[0], point6[0]]
    y_values2 = [point1[1], point2[1], point7[1], point4[1], point8[1], point6[1]]

    plt.plot(x_values, y_values, color='r', linewidth = '3', label=r'Činný výkon $P_{SKUT}$')
    plt.plot(x_values2, y_values2, color='b', ls='--', linewidth = '2', label='Cílový výkon')
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.1), fancybox=True, shadow=True, ncol=5)
    plt.xlabel('Čas (s)')
    plt.ylabel('Výkon (MW)')
    if ceps_product==0:
        plt.title('Test $ mFRR_{12,5A} $ = %s (MW)'%(round(P_cil,2)))
    elif ceps_product==1:
        plt.title('Test $ mFRR_{5A} $ = %s (MW)'%(round(P_cil,2)))

    plt.subplots_adjust(bottom=0.15)

    plt.show()

else:

    fig,ax = plt.subplots()

    plt.xlim(0, 5000)
    plt.ylim(min(P_lim_plus)-0.8, max(P_lim_minus)+0.6)
    point1 = [0 , P_dg]
    point2 = [start_of_plot , P_dg]
    point3 = [start_of_plot+ramp_length, P_dg+P_cil]
    point4 = [start_of_plot+ramp_length+20*60, P_dg+P_cil]
    point5 = [start_of_plot+ramp_length+20*60+ramp_length, P_dg]
    point6 = [start_of_plot+ramp_length+20*60+ramp_length+20*60, P_dg]
    point9 = [start_of_plot+ramp_length+20*60+ramp_length+20*60+start_of_plot, P_dg]

    point7= [start_of_plot, P_dg+P_cil]
    point8=[start_of_plot+ramp_length+20*60, P_dg]

    x_values = [point1[0], point2[0], point3[0], point4[0], point5[0], point6[0], point9[0]]
    y_values = [point1[1], point2[1], point3[1], point4[1], point5[1], point6[1], point9[1]]

    plt.hlines(y=0, xmin=0, xmax=4800, color='#000000')
    plt.text(50,P_dg+0.07, r'$ P_{} = {} (MW) $'.format('{DG}', (round(P_dg,2))), **font3)
    plt.hlines(y=P_dg+P_cil, xmin=0, xmax=4800, color='#000000')
    plt.hlines(y=P_dg, xmin=0, xmax=4800, color='#000000')
    plt.text(50,(P_dg+P_cil)+0.07, r'$ mFRR_{12,5B} = %s (MW) $'%(round(P_cil,2)), **font3)
    plt.vlines(x=point2[0], ymin= min(P_lim_plus)-0.6, ymax = max(P_lim_minus)+0.35, ls = '--', color='#000000')
    plt.vlines(x=point3[0], ymin= min(P_lim_plus)-0.6, ymax = max(P_lim_minus)+0.35, ls = '--', color='#000000')
    plt.vlines(x=point4[0], ymin= min(P_lim_plus)-0.6, ymax = max(P_lim_minus)+0.35, ls = '--', color='#000000')
    plt.vlines(x=point5[0], ymin= min(P_lim_plus)-0.6, ymax = max(P_lim_minus)+0.35, ls = '--', color='#000000')
    plt.vlines(x=point6[0], ymin= min(P_lim_plus)-0.6, ymax = max(P_lim_minus)+0.35, ls = '--', color='#000000')
    plt.xticks([point2[0], point3[0], point4[0], point5[0], point6[0]], [r'$t_{0}$', r'$t_{1}$', r'$t_{2}$', r'$t_{3}$', r'$t_{4}$'])
    plt.yticks([0, round((P_dg+P_cil),2), round(P_dg,2)])
    plt.annotate(text='', xy=(point3[0],max(P_skut+0.6)), xytext=(point2[0],max(P_skut+0.6)), arrowprops=dict(arrowstyle='<->'))
    plt.text(point2[0]+40, max(P_skut+0.7), r'$ t_{AKTmFRR12,5} <= 12,5 (min) $', **font3) 
    plt.annotate(text='', xy=(point4[0],max(P_skut+0.6)), xytext=(point3[0],max(P_skut+0.6)), arrowprops=dict(arrowstyle='<->')) 
    plt.text(point3[0]+400, max(P_skut+0.7), r'$ t_{u} = 20 (min) $', **font3)
    plt.annotate(text='', xy=(point5[0],max(P_skut+0.6)), xytext=(point4[0],max(P_skut+0.6)), arrowprops=dict(arrowstyle='<->'))
    plt.text(point4[0]+40, max(P_skut+0.7), r'$ t_{DEAKTmFRR12,5} <= 12,5 (min) $', **font3)
    plt.annotate(text='', xy=(point6[0],max(P_skut+0.6)), xytext=(point5[0],max(P_skut+0.6)), arrowprops=dict(arrowstyle='<->'))
    plt.text(point5[0]+400, max(P_skut+0.7), r'$ t_{u} = 20 (min) $', **font3)
    plt.text(point5[0]+100, (P_dg)/2-0.04*P_dg, (r'$ t_{0} $ = čas vydání povelu k aktivaci $ mFRR_{12,5} $'), **font3)
    plt.text(point5[0]+100, ((P_dg)/2)-0.08*P_dg, (r'$ t_{1} $ = čas dosažení certifikované hodnoty $ mFRR_{12,5} $'), **font3)
    plt.text(point5[0]+100, ((P_dg)/2)-0.12*P_dg, (r'$ t_{2} $ = čas vydání povelu k deaktivaci $ mFRR_{12,5} $'), **font3) 
    plt.text(point5[0]+100, ((P_dg)/2)-0.16*P_dg, (r'$ t_{3} $ = čas dosažení původního výkonu $ P_{DG} $'), **font3)   
    plt.text(point5[0]+100, ((P_dg)/2)-0.20*P_dg, (r'$ t_{4} $ = čas ukončení testu $ mFRR_{12,5} $'), **font3) 

    x_values2 = [point1[0], point2[0], point7[0], point4[0], point8[0], point6[0]]
    y_values2 = [point1[1], point2[1], point7[1], point4[1], point8[1], point6[1]]

    plt.plot(x_values, y_values, color='r', linewidth = '3', label=r'Činný výkon $P_{SKUT}$')
    plt.plot(x_values2, y_values2, color='b', ls='--', linewidth = '2', label='Cílový výkon')
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.1), fancybox=True, shadow=True, ncol=5)
    plt.xlabel('Čas (s)')
    plt.ylabel('Výkon (MW)')
    plt.title('Test $ mFRR_{12,5B} $ = %s (MW)'%(round(P_cil,2)))
    plt.subplots_adjust(bottom=0.15)

plt.show()

if battery==1:
    fig, ax = plt.subplots()
    plt.xlabel('Čas [s]', **font2)
    plt.ylabel('SoC [%]', **font2)

    plt.xlim(0, excel_source.shape[0]+100)
    plt.ylim(-10, 110)

    plt.plot(Soc, color = 'b', linewidth = '2', label ='SoC (%)')
    plt.hlines(y=20, xmin=0, xmax=4800, color='#000000')
    plt.text(50,21, 'SoC_D', **font3)
    plt.hlines(y=80, xmin=0, xmax=4800, color='#000000')
    plt.text(50,81, 'SoC_H', **font3)
    plt.title('SoC = f(t)')

    # Shrink current axis's height by 10% on the bottom
    box = ax.get_position()
    ax.set_position([box.x0, box.y0 + box.height * 0.05, box.width, box.height * 0.95])

    # Put a legend below current axis
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.1), fancybox=True, shadow=True, ncol=5)

    plt.xticks(np.arange(0, excel_source.shape[0]+100, step=200), **font2)
    plt.yticks(np.arange(-10, 110, step=5), **font2)

    plt.rc('font', **font)

    plt.show()